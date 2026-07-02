from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from uuid import uuid4

from .models import AskResponse


LOG_DIR = Path(__file__).resolve().parents[1] / "logs"
LOG_FILE = LOG_DIR / "analysis.jsonl"


def log_analysis(response: AskResponse, user_ctx: object | None = None) -> str:
    trace_id = uuid4().hex
    user_id: str | None = getattr(user_ctx, "user_id", None)
    tenant_id: str | None = getattr(user_ctx, "tenant_id", None)
    event = {
        "trace_id": trace_id,
        "created_at": datetime.now(UTC).isoformat(),
        "user_id": user_id,
        "tenant_id": tenant_id,
        "question": response.question,
        "intent": response.intent,
        "routing_strategy": response.routing_strategy,
        "llm_provider": response.llm_provider,
        "validation_ok": response.validation.ok,
        "needs_clarification": response.needs_clarification,
        "clarification_option_count": len(response.clarification_options),
        "row_count": len(response.rows),
        "chart_type": response.chart.type if response.chart else None,
        "blocked": response.validation.blocked,
        "orchestration": [step.model_dump() for step in response.orchestration],
    }

    LOG_DIR.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=True) + "\n")

    return trace_id


def recent_events(
    page: int = 1,
    limit: int = 20,
    tenant_id: str | None = None,
    intent: str | None = None,
    validation_ok: bool | None = None,
) -> tuple[list[dict[str, object]], int]:
    """Return (events, total_count) for the given page, newest-first.

    Filters:
      tenant_id     — restrict to this tenant (mandatory for multi-tenant).
      intent        — exact match on the `intent` field (e.g. "margin_trend").
      validation_ok — True = only validated queries; False = only blocked ones.
    """
    if not LOG_FILE.exists():
        return [], 0

    parsed: list[dict[str, object]] = []
    for line in reversed(LOG_FILE.read_text(encoding="utf-8").splitlines()):
        if not line.strip():
            continue
        try:
            ev = json.loads(line)
        except Exception:
            continue
        if tenant_id and ev.get("tenant_id") != tenant_id:
            continue
        if intent and ev.get("intent") != intent:
            continue
        if validation_ok is not None and ev.get("validation_ok") != validation_ok:
            continue
        parsed.append(ev)

    total = len(parsed)
    offset = (page - 1) * limit
    return parsed[offset : offset + limit], total


def export_csv(tenant_id: str | None = None, limit: int = 500) -> str:
    """Return audit events as a CSV string, optionally filtered by tenant_id."""
    import csv
    import io

    FIELDS = [
        "trace_id", "created_at", "user_id", "tenant_id",
        "question", "intent", "routing_strategy", "llm_provider",
        "validation_ok", "needs_clarification", "row_count",
        "chart_type",
    ]

    if not LOG_FILE.exists():
        buf = io.StringIO()
        csv.writer(buf).writerow(FIELDS)
        return buf.getvalue()

    lines = LOG_FILE.read_text(encoding="utf-8").splitlines()[-limit:]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDS, extrasaction="ignore")
    writer.writeheader()

    for line in lines:
        if not line.strip():
            continue
        try:
            ev = json.loads(line)
            if tenant_id and ev.get("tenant_id") != tenant_id:
                continue
            writer.writerow(ev)
        except Exception:
            continue

    return buf.getvalue()


def export_xlsx(tenant_id: str | None = None, limit: int = 500) -> bytes:
    """Return audit events as an xlsx file (bytes), optionally filtered by tenant_id."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    FIELDS = [
        "trace_id", "created_at", "user_id", "tenant_id",
        "question", "intent", "routing_strategy", "llm_provider",
        "validation_ok", "needs_clarification", "row_count", "chart_type",
    ]

    rows: list[dict] = []
    if LOG_FILE.exists():
        for line in LOG_FILE.read_text(encoding="utf-8").splitlines()[-limit:]:
            if not line.strip():
                continue
            try:
                ev = json.loads(line)
                if tenant_id and ev.get("tenant_id") != tenant_id:
                    continue
                rows.append(ev)
            except Exception:
                continue

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit FabriQ"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, field in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(12, len(field) + 2)

    for row_idx, ev in enumerate(rows, start=2):
        for col, field in enumerate(FIELDS, start=1):
            ws.cell(row=row_idx, column=col, value=ev.get(field))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
